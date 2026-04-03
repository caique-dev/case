import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# definindo enderecos dos arquivos utilizados
arquivo_fonte = './Data.xlsx'
arquivo_destino = './consolidated_data.xlsx'

# funções de uso geral
def abrir_ou_criar_planilha(caminho_arquivo: str, nome_aba: str):
    # verifica se o arquivo existe
    if os.path.exists(caminho_arquivo):
        wb = load_workbook(caminho_arquivo)
    else:
        wb = Workbook()
        # remove a aba padrão se for criar um novo arquivo
        if 'Sheet' in wb.sheetnames:
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)

    # verifica se a aba desejada já existe
    if nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
    else:
        ws = wb.create_sheet(title=nome_aba)

    return wb, ws

def inserir_df_na_aba(ws: any, df: pd.DataFrame, incluir_header: bool = True, incluir_index: bool = False):
    """
    ws: worksheet do openpyxl
    df: pandas DataFrame
    incluir_header (bool): inclui nomes das colunas
    incluir_index (bool): inclui índice do DataFrame
    """
    # limpando possiveis dados existentes na aba
    ws.delete_rows(1, ws.max_row)

    # iterando sobre as linhas e colunas do df e inserindo os valores célula a célula na planilha
    for r_idx, row in enumerate( # linhas
        dataframe_to_rows(df, index=incluir_index, header=incluir_header),
        start=1 # index da linha onde os dados começarão a ser inseridos, o padrão do enumerate é 0, diferindo da indexação do Excel que começa em 1
    ):
        for c_idx, value in enumerate(row, start=1): # colunas
            ws.cell(row=r_idx, column=c_idx, value=value)

# teste 1: estruturação de dados
def teste_1():
    # etapa A: consolidando os dados brutos em uma única aba
    ## utilizei o wide format na consolidação dos dados, onde cada papel tem uma coluna com seus preços, pois facilita a visualização e simplifica a tarefa, pois mantém o formato original dos dados
    def etapa_a(param_fonte: str, param_destino: str, param_aba_destino: str):
        # transformando o arquivo fonte em um dict onde as chaves são os nomes das abas e os valores são os dataframes correspondentes a cada aba
        todas_abas = pd.read_excel(param_fonte, sheet_name=None) # sheet_name=None retorna um dicionário onde as chaves são os nomes das abas e os valores são os dataframes correspondentes a cada aba
        aux = []

        # gerando dfs com os dados de cada aba, mantendo a aba "Data" e alterando a coluna "Preço" para o nome da aba correspondente
        for nome_aba, df in todas_abas.items():
            df = df.rename(columns={'Preço': f'{nome_aba}'})
            aux.append(df)

        # pegando os dados da primeira aba para iniciar a consolidação
        infos_consolidadas = aux[0] 
        for df in aux[1:]:
            # unindo os dados dados das abas seguintes, utilizando a coluna "Data" como chave de junção
            infos_consolidadas = pd.merge(infos_consolidadas, df, on='Data', how='inner')

        # guardando a consolidação no excel de destino
        wb, ws = abrir_ou_criar_planilha(param_destino, param_aba_destino)
        inserir_df_na_aba(ws, infos_consolidadas)
        wb.save(param_destino)

    # etapa B: calculando o retorno diário de cada papel
    ## Nesta etapa achei necessário alterar a tabela para o formato longo, onde cada linha representa o preço de um papel em uma data, tanto para me adequar melhor às exigências do teste 2A, que solicita a adição de apenas uma nova coluna, quanto para que a tabela fique menor, com apenas 3 colunas em vez de 11
    def etapa_b(param_fonte:str, param_aba_destino: str = ''):
        # lendo o arquivo consolidado
        df_precos = pd.read_excel(param_fonte) 

        # definindo a coluna "Data" como índice para facilitar o cálculo do retorno diário
        df_precos = df_precos.set_index('Data') 
        
        # calculando os retornos diários dos papéis
        df_retornos = df_precos.pct_change()

        # calculando o retorno diário médio de cada papel e ordenando a serie de forma decrescente
        retorno_medio = df_retornos.mean().sort_values(ascending=False)
        
        # pegando os nomes dos 5 papéis com maior retorno médio
        top_5 = retorno_medio.head(5).index.tolist()

        # filtrando o top 5 no df de precos
        df_final = df_precos[top_5]

        # resetando o índice para que a coluna "Data" volte a ser uma coluna normal, necessário para alterar o formato do df e para facilitar a exportação para o excel
        df_final = df_final.reset_index()

        # alterando o formato do df para long
        df_final = df_final.melt(
            id_vars='Data',
            var_name='Papel',
            value_name='Preço'
        )

        # salvando o df final na aba top_5
        wb, ws = abrir_ou_criar_planilha(param_fonte, param_aba_destino)
        inserir_df_na_aba(ws, df_final)
        wb.save(param_fonte)
    
    # executando as etapas
    etapa_a(arquivo_fonte, arquivo_destino, param_aba_destino='consolidated_data')
    etapa_b(arquivo_destino, param_aba_destino='top_5')

teste_1()